#!/usr/bin/python3
import openpyxl
import sys

'''
  To do:
    Can't calculate the result
'''

#wb = openpyxl.Workbook()
#wb.create_sheet('test_case')
#wb.save('cases.xlsx')
#sheet = wb.active
#sheet.sheet_properties.tabColor="1072BA"

def writeData2execl(kbpsLi,psnrYLi,filePath):
  wb = openpyxl.load_workbook(filePath,)
  sh = wb['C4']
  #write kbpsLi
  i = 3
  for kbps in kbpsLi:
    ce = sh.cell(row = i,column = 20)
    ce.value = kbps
    i = i + 1

  i = 3
  for psnr in psnrYLi:
    ce = sh.cell(row = i,column = 21)
    ce.value = psnr
    i = i + 1

  wb.save('AVS.xlsx')
  wb.close()
  
def getDataLi(filePath):  
  li = []
  with open(filePath,"r") as f:
    line = f.readline()
    while(line):
      li.append(line)
      line = f.readline()
      #print(line,end='')
    return li    

kbpsLi = getDataLi(sys.argv[1])
psnrYLi= getDataLi(sys.argv[2]) 
writeData2execl(kbpsLi,psnrYLi,sys.argv[3])
