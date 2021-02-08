#!/usr/bin/python3
import openpyxl
from openpyxl.styles import Font, colors, Alignment
import sys

wb = openpyxl.Workbook()
#wb.create_sheet('test_case')
#wb.save('cases.xlsx')
#wb = openpyxl.load_workbook('cases.xlsx')

#sh = wb['test_case']
ws = wb.active
#sheet.sheet_properties.tabColor="1072BA"
font_set = Font(name='Arial', size=20, italic=True, color=colors.BLUE)
#ws.column_dimensions['A'].height = 30 
i = 0
with open(sys.argv[1],"r") as f:
    line = f.readline()
    line = line.strip("\n") 
    while(line):
        #print(line,end="")
        #sheet["A%d" % (i+1)].value = str(line)
        
        #ws['A%d' % (i+1) ].font = font_set
        line = line.split()
        ws['A%d' % (i+1) ] = line[0]
        ws['B%d' % (i+1) ] = line[1]
        
        i=i+1
        line = f.readline()
        line = line.strip("\n") 



wb.save('cases.xlsx')
wb.close()
